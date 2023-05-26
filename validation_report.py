import pyodbc
import openpyxl
import datetime
from openpyxl.utils import get_column_letter

#   VARIABLES               ***********************************************************************

src_server = "[COM-DLKDWD01.SITUSAMC.COM]"
# src_server = "[COM-DWDBU01.SITUSAMC.COM]"
# src_server = "[COM-DWAGP01.SITUSAMC.COM]"

src_db = "[DW_PROPERTY_VALUATION]"
meta_db = "[DW_META]"

save_path = "C:\\Users\\jasonwalker\\OneDrive - SitusAMC\\Documents\\project\\VALIDATION_REPORT\\FILE_OUTPUT\\"
# save_path = ""



sql_validations = """SELECT 
[DATA_SET]
,[SHEET_NAME]
,[RESOLUTION]
,[META_INSERT]
,[META_SELECT]
,[META_WHERE]
FROM [stage].[META_VALIDATION]
WHERE [IS_INACTIVE] IS NULL
AND [IS_PRIME] IS NOT NULL
--AND [META_VALIDATION_KEY] IN (10000077,10000078)
ORDER BY [SHEET_ORDER]"""

solution_font = openpyxl.styles.Font(color="0070C0", bold=True, size=14)
header_font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=14)
column_font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)


substitutions = [
    ("AND F.[CLIENT_KEY] = @CLIENT_KEY", ""),
    ("WHERE [CLIENT_KEY] = @CLIENT_KEY", ""),
    ("WHERE F.[CLIENT_KEY] = @CLIENT_KEY", ""),
    ("WHERE F.[CLIENT_KEY] = @CLIENT_KEY)", ""),
    ("AND [CLIENT_KEY] = @CLIENT_KEY", ""),
    ("AND S.[CLIENT_KEY] = @CLIENT_KEY", ""),
    ("WHERE [DW_DIMENSIONS].[stage].[fn_CLIENT_KEY]('CLIENT', 120010, NULL, NULL, B.[ClientName]) = @CLIENT_KEY", ""),
    ("WHERE [DW_DIMENSIONS].[stage].[fn_CLIENT_KEY]('CLIENT', 120016, NULL, NULL, S.[ClientName]) = @CLIENT_KEY", ""),
    ("WHERE [DW_DIMENSIONS].[stage].[fn_CLIENT_KEY]('CLIENT', 120010, NULL, NULL, F.[ClientName]) = @CLIENT_KEY", ""),
    ("WHERE [DW_DIMENSIONS].[stage].[fn_CLIENT_KEY]('CLIENT', 120010, NULL, NULL, [ClientName]) = @CLIENT_KEY", ""),
    ("AND [DW_DIMENSIONS].[stage].[fn_CLIENT_KEY]('CLIENT', 120010, NULL, NULL, F.[ClientName]) = @CLIENT_KEY", ""),
    ("AND [DW_DIMENSIONS].[stage].[fn_CLIENT_KEY]('CLIENT',110003,NULL,NULL,[CLIENTNAME]) = @CLIENT_KEY", "")
]


#   DATABASE CONNECTIONS    ***********************************************************************

# if src_server == "[COM-DLKDWD01.SITUSAMC.COM]":
#     pwdone = "BPcpfhSx0t2669St2jDX"
# else:
pwdone = "FZ9H2Pcg=z%-cf?$"

conn01 = pyodbc.connect(
    "Driver={SQL Server};"
    f"Server={src_server[1:-1]};"
    f"Database={src_db[1:-1]};"
    "UID=python;"
    f"PWD={pwdone};"
    , autocommit=True)

conn02 = pyodbc.connect(
    "Driver={SQL Server};"
    f"Server={src_server[1:-1]};"
    f"Database={meta_db[1:-1]};"
    "UID=python;"
    f"PWD={pwdone};"
    , autocommit=True)


#   FUNCTIONS    **********************************************************************************

def filename(client_name):
    client_name = client_name.replace(" ", "_")
    year = str(datetime.date.today().year)
    if datetime.date.today().month <= 9:
        month = "0" + str(datetime.date.today().month)
    else:
        month = str(datetime.date.today().month)
    if datetime.date.today().day <= 9:
        day = "0" + str(datetime.date.today().day)
    else:
        day = str(datetime.date.today().day)
    return client_name + "_" + year + month + day + ".xlsx"


def validation_log(data_set, val_name, result, run_start, run_stop, run_time, row_count):
    log_insert = f"insert into [stage].[META_VALIDATION_LOG] values({run_batch[0][0]}, '{data_set}', '{val_name}', '{result}', '{str(run_start)[:22]}', '{str(run_stop)[:22]}', '{run_time}', {row_count}, GETDATE(), USER_ID(), NULL, NULL)"
    cursor2.execute(log_insert)
    return


def isnull(x, y):
    if x != x:
        reply = y
    elif x is None:
        reply = y
    else:
        reply = x
    return reply


#   MAIN APPLICATION        ***********************************************************************
t1a = datetime.datetime.now()

cursor = conn01.cursor()
cursor2 = conn02.cursor()

run_batch = list(cursor2.execute("SELECT MAX([RUN_BATCH])+1 FROM [DW_META].[stage].[META_VALIDATION_LOG]"))
if run_batch[0][0] is None:
    run_batch[0][0] = 100000


validation_file = filename("VALIDATIONS_PRIME")
workbook = openpyxl.Workbook()

#    loops through each validation (sheet)
validation_summary = []
validation_list = list(cursor2.execute(sql_validations))
for index2, validation in enumerate(validation_list):
    t2a = datetime.datetime.now()
    validation_name = validation[1]
    validation_sql = validation[4] + " " + isnull(validation[5], "")

    for search, replacement in substitutions:
        validation_sql = validation_sql.replace(search, replacement)

    try:
        # print(validation_sql)
        validation_results = list(cursor.execute(validation_sql))
        rowcount = len(validation_results)
        t2b = datetime.datetime.now()
        validation_log(validation[0], validation[1], "SUCCESSFUL", t2a, t2b, t2b - t2a, rowcount)
        print(f"    SUCCESSFUL:     {validation_name.ljust(35)}  run time: {str(t2b - t2a).rjust(16)}       row count: {str(rowcount).rjust(15)}")
    except:
        t2c = datetime.datetime.now()
        validation_log(validation[0], validation[1], "FAILED", t2a, t2c, t2c - t2a, 0)
        print("\n" * 2)
        print(f"    FAILED:         {validation_name.ljust(35)}  run time: {str(t2c - t2a).rjust(16)}")
        print("\n")
        print(validation_sql)
        print("\n" * 4)
        continue

    sheet = workbook.create_sheet(validation[1])
    sheet["A1"] = f"=HYPERLINK(\"#\'VALIDATION SUMMARY\'!A1\", \"Home\")"
    sheet["B1"] = validation[2]

    #   populate column headers and start sizing column widths
    column_widths = []
    for indx, column in enumerate(cursor.description, 1):
        sheet.cell(row=2, column=indx, value=column[0])
        column_widths += [len(str(column[0]))]

    #   populate cells in sheet and determine column width
    for r_idx, row in enumerate(validation_results, 0):
        for c_idx, value in enumerate(row, 0):
            sheet.cell(row=3 + r_idx, column=1 + c_idx, value=row[c_idx])
            #   calculate max column width
            if len(column_widths) > c_idx:
                if len(str(row[c_idx])) > column_widths[c_idx]:
                    column_widths[c_idx] = len(str(row[c_idx]))
            else:
                column_widths += [len(str(row[c_idx]))]
    validation_summary.append((validation[0], validation[1], rowcount))
    # print(validation[0].ljust(30), "row count:  ", len(validation_results))

    #   APPLYING STYLES
    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = column_width + 5
        sheet.cell(row=2, column=i).fill = openpyxl.styles.PatternFill(start_color="2F9DFA", end_color="2F9DFA", fill_type="solid")
        sheet.cell(2, i).font = column_font

    sheet["A1"].font = solution_font
    sheet["B1"].font = solution_font
    row_one_height = sheet.row_dimensions[1]
    row_one_height.height = 30
    row_one_height = sheet.row_dimensions[2]
    row_one_height.height = 20
    sheet.freeze_panes = "A3"


# create validation summary tab
title_date = datetime.datetime.now()
val_sum_sheet = workbook["Sheet"]
val_sum_sheet.title = "VALIDATION SUMMARY"
val_sum_sheet.merge_cells("A1:C1")

val_sum_title = val_sum_sheet["A1"]
val_sum_title.value = f"""Validation Results Summary: {src_server}  {title_date.strftime("%Y")}.{title_date.strftime("%m")}.{title_date.strftime("%d")}"""
val_sum_title.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
val_sum_title.font = header_font
val_sum_title.fill = openpyxl.styles.GradientFill(stop=("091D44", "2F3CF3"))

val_sum_sheet.row_dimensions[1].height = 30
val_sum_sheet.column_dimensions["A"].width = 30
val_sum_sheet.column_dimensions["B"].width = 60
val_sum_sheet.column_dimensions["C"].width = 10
val_sum_sheet.freeze_panes = "A3"
val_sum_sheet.sheet_view.showGridLines = False


for r_idx, row in enumerate(validation_summary, 0):
    val_sum_sheet.cell(row=3 + r_idx, column=1, value=row[0])
    val_sum_sheet.cell(row=3 + r_idx, column=2, value=row[1])
    val_sum_sheet.cell(row=3 + r_idx, column=3).value = f"=HYPERLINK(\"#\'{row[1]}\'!A1\", \"{row[2]}\")"
    val_sum_sheet.cell(row=3 + r_idx, column=3).alignment = openpyxl.styles.Alignment(horizontal="right")

workbook.save(save_path + validation_file)


#   PERFORMANCE TIMES       ***********************************************************************

t1b = datetime.datetime.now()

print("\n" * 2)
print("total start:".ljust(30), t1a)
print("total end:".ljust(30), t1b)
print("total run time:".ljust(30), t1b - t1a)
